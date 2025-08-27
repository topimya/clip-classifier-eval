from transformers import CLIPProcessor, CLIPModel
import cv2
import os
import torch
import numpy as np
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tools import parse_args

device = 'cuda'
start_time = time.perf_counter()
model = CLIPModel.from_pretrained("openai/clip-vit-base-patch32" ).to(device)
processor = CLIPProcessor.from_pretrained("openai/clip-vit-base-patch32")
load_time = time.perf_counter() - start_time
eps = 1e-13

print(f"Время загрузки модели и процессора: {load_time:.2f} секунд")

classes_names = ['']

def read_classes(file_path, flag):
    label_to_index = {}
    class_names = []
    with open(file_path, "r") as f:
        for idx, line in enumerate(f):
            if flag:
                synonyms = [word.strip() for word in line.strip().split(',')]
            else:
                synonyms = [line.strip()]
            for synonym in synonyms:
                class_names.append(synonym)
                label_to_index[synonym] = idx
    return class_names, label_to_index


def clip_probs(images, labels):
    inputs = processor(text=labels, images=images, return_tensors="pt", padding=True).to(device)
    with torch.no_grad():
        start_time = time.perf_counter()
        outputs = model(**inputs)
    probs = outputs.logits_per_image.softmax(dim=1).cpu().numpy()
    inference_time = time.perf_counter() - start_time

    return probs, inference_time

def save_img(image, clip_idx, output_dir, labels, max_prob):
    font = cv2.FONT_HERSHEY_SIMPLEX # шрифт
    color = (255, 255, 255) # цвет
    thickness = 2 # толщина линии 

    name = os.path.basename(image)
    text = f"{labels[clip_idx]}, {max_prob:.2f}%"
    file = open(image, "rb")
    bytes = bytearray(file.read())
    numpyarray = np.asarray(bytes, dtype=np.uint8)
    image = cv2.imdecode(numpyarray, cv2.IMREAD_UNCHANGED)

    h, w = image.shape[:2]

    min_size = 300
    top = bottom = max((min_size - h) // 2, 0)
    left = right = max((min_size - w) // 2, 0)
    image = cv2.copyMakeBorder(image, top, bottom, left, right, cv2.BORDER_CONSTANT, value=(0, 0, 0))
    h, w = image.shape[:2]

    font_scale = 0.8 * (min(w, h) / 500) # размер шрифта
    (text_w, text_h), baseline = cv2.getTextSize(text, font, font_scale, thickness)
    overlay = image.copy()
    padding = 5
    position = (padding, text_h + padding)
    cv2.rectangle(overlay, (position[0] - 5, position[1] - text_h - 5), (position[0] + text_w + 5, position[1] + baseline + 5), (0, 0, 0), -1)
    alpha = 0.5
    cv2.addWeighted(overlay, alpha, image, 1 - alpha, 0, image)
    cv2.putText(image, text, position, font, font_scale, color, thickness)
    save_name = os.path.join(output_dir, name)
            
    _, buffer = cv2.imencode('.jpg', image)
    with open(save_name, 'wb') as f:
        f.write(buffer)
    return 

def metricks(metrics_detailed, percent_thresholds, name_img, probs_all, root, output_dir, labels, label_to_index, metricks_class):
    for i, name in enumerate(name_img):
        j = int(name.split('_')[-2])
        probs = probs_all[i]
        clip_idx = np.argmax(probs)
        max_prob = probs[clip_idx] * 100
        predicted_label = labels[clip_idx]
        predicted_index = label_to_index[predicted_label]
        class_clip = classes_names[predicted_index]
        true_indx = j

        image = os.path.join(root, name)
        save_img(image, clip_idx, output_dir, labels, max_prob)
        
        metrics_detailed.loc[(slice(None), [classes_names[j], 'sum']), 'N'] += 1
        metricks_class.loc[classes_names[true_indx], class_clip] += 1
        mask = np.array(percent_thresholds) <= max_prob
        if predicted_index == true_indx:
            metrics_detailed.loc[(np.array(percent_thresholds)[mask], [class_clip, 'sum']), 'TP'] += 1
        else:
            metrics_detailed.loc[(np.array(percent_thresholds)[mask], [class_clip, 'sum']), 'FP'] += 1
    metrics_detailed['Precision'] = metrics_detailed['TP'] / (metrics_detailed['TP'] + metrics_detailed['FP'] + eps)
    metrics_detailed['Recall'] = metrics_detailed['TP'] / (metrics_detailed['N'] + eps)
    metrics_detailed['F1'] = 2 * (metrics_detailed['Precision'] * metrics_detailed['Recall']) / (metrics_detailed['Precision'] + metrics_detailed['Recall'] + eps)
    return metrics_detailed, metricks_class

def clip_main(dir, res_dir):
    metrics_names = ['TP', 'FP', 'N', 'Precision', 'Recall', 'F1']
    percent_thresholds = list(range(5, 100, 5))
    indexes = [percent_thresholds, classes_names + ['sum']]
    multi_index = pd.MultiIndex.from_product(indexes, names=['%', 'class'])

    all_metrics = pd.DataFrame(data={name: 0 for name in metrics_names}, index=multi_index)

    metricks_class = pd.DataFrame(data={name: 0 for name in classes_names}, index=classes_names)

    clip_res_dir = os.path.join(res_dir, f'clip')
    os.makedirs(clip_res_dir, exist_ok=True)

    labels, label_to_index = read_classes(os.path.join(res_dir, 'classes.txt'), FLAG)
    with open(os.path.join(res_dir, 'labels.txt'), "w") as f:
        f.write('. '.join(labels))
    with pd.ExcelWriter(os.path.join(res_dir,f'{os.path.basename(res_dir)}.xlsx'), engine='openpyxl') as writer:
        for root, _, files in os.walk(dir):
            img = any(file.lower().endswith(('.png', '.jpg', '.jpeg')) for file in files)
            if img:
                print(f'Обработка папка: {os.path.basename(root)}')

                metrics_detailed = pd.DataFrame(data={name: 0 for name in metrics_names}, index=multi_index)

                output_dir = os.path.join(clip_res_dir, f'{os.path.basename(root)}')
                os.makedirs(output_dir, exist_ok=True)

                images = []
                name_img = []
                for img in os.listdir(root):
                    name_img.append(img)
                    img = os.path.join(root, img)
                    file = open(img, "rb")
                    bytes = bytearray(file.read())
                    numpyarray = np.asarray(bytes, dtype=np.uint8)
                    img = cv2.imdecode(numpyarray, cv2.IMREAD_UNCHANGED)
                    rgb_image = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                    images.append(rgb_image)

                probs_all, inf_time = clip_probs(images, labels)
                print(f"Время для {len(images)} изображений: {inf_time:.3f} секунд")

                metrics_detailed, metricks_class = metricks(metrics_detailed, percent_thresholds, name_img, probs_all, root, output_dir, labels, label_to_index, metricks_class)
                metrics_detailed.to_excel(excel_writer=writer, sheet_name=os.path.basename(root))
                all_metrics += metrics_detailed

        all_metrics['Precision'] = all_metrics['TP'] / (all_metrics['TP'] + all_metrics['FP'] + eps)
        all_metrics['Recall'] = all_metrics['TP'] / (all_metrics['N'] + eps)
        all_metrics['F1'] = 2 * (all_metrics['Precision'] * all_metrics['Recall']) / (all_metrics['Precision'] + all_metrics['Recall'] + eps)
        all_metrics.to_excel(excel_writer=writer, sheet_name='all')
        
        cls_dir = os.path.join(res_dir, f'classification_report_{os.path.basename(res_dir)}.xlsx')
        metricks_class.to_excel(cls_dir)
        wb = load_workbook(cls_dir)
        ws = wb.active
        fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # светло-зелёны
        for i in range(len(classes_names)):
            cell = ws.cell(row=2+i, column=2+i)
            cell.fill = fill
        wb.save(cls_dir)
        
    return 

if __name__ == '__main__':

